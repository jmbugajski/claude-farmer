"""
parse_ecowitt.py
================
Read EcoWitt GW1200 daily-log .xlsx exports from the inputs/ folder and return a
single, clean, time-sorted list of readings.

EcoWitt "all_..." exports have a two-row header:

    row 0 (group):  Time | Indoor | ... | Tomato Probe | ... | Pepper Probe | ... | WFC01-... | ...
    row 1 (sub):         | Temperature(F) | ... | Soil Moisture(%) | AD | Soil Moisture(%) | AD | Water Total(L) | ...

Columns are located by matching (group, sub) names from config rather than by a
fixed index, so re-ordered or renamed exports keep working as long as the probe
group names in config.json match what you set on the EcoWitt console.

Missing sensor values are exported as the literal string "-"; those become None.
"""

from __future__ import annotations

import glob
import os
import re
from datetime import datetime
from typing import Optional

import openpyxl

DATA_SHEET = "result_list"


def _norm(x) -> str:
    """Lowercase and strip non-alphanumerics, so 'WFC01-00003D29' and
    '[WFC01] Water Flow' both normalize to a string containing 'wfc01'."""
    return re.sub(r"[^a-z0-9]", "", str(x).lower())


def _num(v) -> Optional[float]:
    """Coerce a cell to float, treating '-'/blank/text as missing (None)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "--"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_time(v) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _find_column(group_row, sub_row, group_name, sub_name) -> Optional[int]:
    """
    Return the column index whose group header == group_name (prefix match ok)
    and sub header == sub_name. Group headers are merged cells, so the group
    label only appears on the first column of the group and the following
    columns carry None — we forward-fill it.
    """
    filled_group = []
    last = None
    for g in group_row:
        if g is not None and str(g).strip() != "":
            last = str(g).strip()
        filled_group.append(last)

    gn = _norm(group_name)
    for i, (g, s) in enumerate(zip(filled_group, sub_row)):
        if g is None or s is None:
            continue
        g, s = str(g).strip(), str(s).strip()
        # Group match tolerates EcoWitt header renames (e.g. 'WFC01-00003D29'
        # -> '[WFC01] Water Flow'): exact, prefix, or normalized-substring.
        group_ok = g == group_name or g.startswith(group_name) or gn in _norm(g)
        if s == sub_name and group_ok:
            return i
    return None


def _find_voltage(group_row, sub_row, ch: str) -> Optional[int]:
    """
    Locate a per-channel sensor voltage column. EcoWitt has renamed these
    across firmware versions, e.g. 'Soil Moisture Sensor CH1(V)' ->
    '[CH1] Tomato Soil Sensor(V)', so match on the channel tag + '(V)'.
    """
    for i, s in enumerate(sub_row):
        if s is None:
            continue
        s = str(s).strip()
        if ch.lower() in s.lower() and "(v)" in s.lower():
            return i
    return None


def load_readings(inputs_dir: str, config: dict) -> list[dict]:
    """
    Load every .xlsx in inputs_dir and return a de-duplicated, time-sorted list
    of {"dt": datetime, "tom": float|None, "pep": float|None, "water": float|None}.
    """
    cols = config["columns"]
    tom_group = config["probes"]["tomato"]["group"]
    pep_group = config["probes"]["pepper"]["group"]
    water_prefix = config["water"]["group_prefix"]
    sm_sub = cols["soil_moisture_sub"]
    water_sub = cols["water_total_sub"]

    files = sorted(glob.glob(os.path.join(inputs_dir, "*.xlsx")))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {inputs_dir}")

    by_dt: dict[datetime, dict] = {}

    for path in files:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[DATA_SHEET] if DATA_SHEET in wb.sheetnames else wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 3:
            continue

        group_row, sub_row = rows[0], rows[1]
        i_tom = _find_column(group_row, sub_row, tom_group, sm_sub)
        i_pep = _find_column(group_row, sub_row, pep_group, sm_sub)
        i_water = _find_column(group_row, sub_row, water_prefix, water_sub)
        # Diagnostic channels: raw AD count per probe + per-channel sensor
        # voltage. Used for sensor-health checks (a battery swap or chemistry
        # change shifts the derived %, and a failing/uncoupled probe shows a
        # collapsing daily AD range) — see analyze._sensor_health.
        i_tom_ad = _find_column(group_row, sub_row, tom_group, "AD")
        i_pep_ad = _find_column(group_row, sub_row, pep_group, "AD")
        i_v1 = _find_voltage(group_row, sub_row, "CH1")
        i_v2 = _find_voltage(group_row, sub_row, "CH2")

        for r in rows[2:]:
            dt = _parse_time(r[0])
            if dt is None:
                continue
            def cell(i):
                return _num(r[i]) if i is not None else None
            by_dt[dt] = {
                "dt": dt,
                "tom": cell(i_tom),
                "pep": cell(i_pep),
                "water": cell(i_water),
                "tom_ad": cell(i_tom_ad),
                "pep_ad": cell(i_pep_ad),
                "v_tom": cell(i_v1),
                "v_pep": cell(i_v2),
            }

    return [by_dt[k] for k in sorted(by_dt)]


if __name__ == "__main__":
    import json
    import sys

    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    cfg = json.load(open(os.path.join(here, "config.json")))
    recs = load_readings(os.path.join(here, "inputs"), cfg)
    print(f"Loaded {len(recs)} readings")
    print("first:", recs[0])
    print("last :", recs[-1])
